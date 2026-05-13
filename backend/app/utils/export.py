import io
import random
from datetime import datetime, timedelta, timezone
from html import escape
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
from collections import OrderedDict

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    KeepTogether,
    HRFlowable,
)

from .chart_generator import generate_plc_graph_images

try:
    IST_TIMEZONE = ZoneInfo("Asia/Kolkata")
except ZoneInfoNotFoundError:
    IST_TIMEZONE = timezone(timedelta(hours=5, minutes=30), name="IST")

# Normalize rows.

def _normalize_rows(rows: list) -> list[list]:
    normalized: list[list] = []
    for row in rows or []:
        if isinstance(row, (list, tuple)):
            normalized.append(list(row))
        else:
            normalized.append([row])
    return normalized

# Handle excel fill.

def _excel_fill(hex_color: str) -> PatternFill:
    color = str(hex_color or "").strip().lstrip("#")
    return PatternFill(fill_type="solid", start_color=color, end_color=color)

# Handle excel autofit columns.

def _excel_autofit_columns(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        lengths = []
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            lengths.append(len(value))
        width = max(lengths) + 2 if lengths else min_width
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(min_width, min(width, max_width))

# Handle excel apply table style.

def _excel_apply_table_style(
    ws,
    *,
    header_row: int,
    first_data_row: int,
    last_data_row: int,
    col_count: int,
    header_fill: PatternFill,
    alt_row_fill: PatternFill,
    header_text_color: str = "FFFFFF",
    border_color: str = "D1D5DB",
) -> None:
    if col_count <= 0:
        return

    thin = Side(border_style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, col_count + 1):
        head = ws.cell(row=header_row, column=col)
        head.fill = header_fill
        head.font = Font(bold=True, color=header_text_color)
        head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        head.border = border

    for row_idx in range(first_data_row, last_data_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            if (row_idx - first_data_row) % 2 == 1:
                cell.fill = alt_row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

# Handle generated at text.

def _generated_at_text() -> str:
    return datetime.now(IST_TIMEZONE).strftime("%d %b %Y %I:%M %p")

# Handle export dispatch report pdf.

def export_dispatch_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
    date_column_index: int = 0,
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    DATE_BG     = HexColor("#E8EEF7")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 1.6 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # ── Group rows by date ────────────────────────────────────────────────────
    grouped: OrderedDict[str, list] = OrderedDict()
    for row in normalized_rows:
        date_key = str(row[date_column_index]).strip()
        try:
            display_date = datetime.strptime(date_key, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            display_date = date_key
        grouped.setdefault(display_date, []).append(row)
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "DISPATCH REPORT")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Records: {len(normalized_rows)}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Dispatch Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_cell_style = ParagraphStyle(
        "DRHeaderCell", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    data_cell_style = ParagraphStyle(
        "DRDataCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    summary_label_style = ParagraphStyle(
        "DRSummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "DRSummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    date_heading_style = ParagraphStyle(
        "DRDateHeading", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", leading=13,
    )
    date_sub_style = ParagraphStyle(
        "DRDateSub", parent=styles["Normal"],
        fontSize=7.5, textColor=GREY_MID, leading=11,
    )
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("DISPATCH SUMMARY", summary_label_style),
        Paragraph(f"Total Records: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Days: <b>{len(grouped)}</b>", summary_value_style),
        Paragraph(f"Generated: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.28, 0.22, 0.14, 0.36)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Helper: build date heading banner ─────────────────────────────────────

    def build_date_banner(display_date: str, record_count: int) -> Table:
        banner_data = [[
            Paragraph(display_date, date_heading_style),
            Paragraph(f"{record_count} record{'s' if record_count != 1 else ''}", date_sub_style),
        ]]
        banner = Table(banner_data, colWidths=[AVAIL_W * 0.75, AVAIL_W * 0.25])
        banner.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), DATE_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, NAVY),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RULE),
            ("ALIGN",         (1, 0), (1, 0),   "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return banner
 
    # ── Assemble story ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
 
    story = [
        summary_tbl,
        Spacer(1, 5 * mm),
    ]
 
    for i, (display_date, day_rows) in enumerate(grouped.items()):
        banner = build_date_banner(display_date, len(day_rows))
        
        # Add banner
        story.append(banner)
        story.append(Spacer(1, 1.5 * mm))
        
        # Build combined table: header + all data rows
        col_w = AVAIL_W / len(headers)
        combined_data = [[Paragraph(str(h), header_cell_style) for h in headers]]
        
        for row in day_rows:
            combined_data.append([Paragraph(str(cell), data_cell_style) for cell in row])
        
        combined_tbl = Table(combined_data, colWidths=[col_w] * len(headers), repeatRows=1)
        combined_tbl.setStyle(TableStyle([
            # Header row styling
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1, 0),  8),
            ("TOPPADDING",     (0, 0), (-1, 0),  9),
            ("BOTTOMPADDING",  (0, 0), (-1, 0),  9),
            ("VALIGN",         (0, 0), (-1, 0),  "MIDDLE"),
            ("ALIGN",          (0, 0), (-1, 0),  "CENTER"),
            # Data rows
            ("TOPPADDING",     (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING",  (0, 1), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("TEXTCOLOR",      (0, 1), (-1, -1), GREY_DARK),
            ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",       (0, 1), (-1, -1), 8),
            ("VALIGN",         (0, 1), (-1, -1), "MIDDLE"),
            # Shared
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
            # Lines
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]))
        
        story.append(combined_tbl)
        
        # Gap between day sections (not after the last one)
        if i < len(grouped) - 1:
            story.append(Spacer(1, 6 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()
 
# Handle export dispatch entry report pdf.

def export_dispatch_entry_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    TOTAL_BG    = HexColor("#EEF3FA")   # subtle blue tint for TOTAL row
    TOTAL_FG    = HexColor("#0D2545")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 1.6 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # Separate data rows from TOTAL row (last row if first cell is "TOTAL")
    data_rows  = normalized_rows
    total_row  = None
    if normalized_rows and str(normalized_rows[-1][0]).strip().upper() == "TOTAL":
        data_rows = normalized_rows[:-1]
        total_row = normalized_rows[-1]
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "DISPATCH ENTRY REPORT")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Rows: {len(data_rows)}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Dispatch Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_cell_style = ParagraphStyle(
        "DEHeaderCell", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    data_cell_style = ParagraphStyle(
        "DEDataCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    total_label_style = ParagraphStyle(
        "DETotalLabel", parent=styles["Normal"],
        fontSize=8, textColor=TOTAL_FG, fontName="Helvetica-Bold", leading=11,
    )
    total_value_style = ParagraphStyle(
        "DETotalValue", parent=styles["Normal"],
        fontSize=8, textColor=TOTAL_FG, fontName="Helvetica-Bold", leading=11,
    )
    summary_label_style = ParagraphStyle(
        "DESummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "DESummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    section_style = ParagraphStyle(
        "DESection", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4,
    )
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("DISPATCH ENTRY SUMMARY", summary_label_style),
        Paragraph(f"Total Rows: <b>{len(data_rows)}</b>", summary_value_style),
        Paragraph(f"Columns: <b>{len(headers)}</b>", summary_value_style),
        Paragraph(f"Date: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.32, 0.16, 0.16, 0.36)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Data table ────────────────────────────────────────────────────────────
    col_w = AVAIL_W / max(len(headers), 1)
 
    # Header row
    table_data = [[Paragraph(str(h), header_cell_style) for h in headers]]
 
    # Data rows
    for row in data_rows:
        table_data.append([Paragraph(str(cell), data_cell_style) for cell in row])
 
    # TOTAL row pinned at bottom with distinct styling
    total_row_index = None
    if total_row is not None:
        total_row_index = len(table_data)
        table_data.append([
            Paragraph(str(cell), total_label_style if i == 0 else total_value_style)
            for i, cell in enumerate(total_row)
        ])
 
    data_tbl = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
 
    style_cmds = [
        ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
        ("TOPPADDING",     (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING",  (0, 0), (-1, 0),  9),
        ("TOPPADDING",     (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING",  (0, 1), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
        ("LEFTPADDING",    (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
        ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
        ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
        ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
    ]
 
    # Override TOTAL row styling
    if total_row_index is not None:
        style_cmds += [
            ("BACKGROUND", (0, total_row_index), (-1, total_row_index), TOTAL_BG),
            ("LINEABOVE",  (0, total_row_index), (-1, total_row_index), 1.2, NAVY),
            ("LINEBELOW",  (0, total_row_index), (-1, total_row_index), 1.2, NAVY),
        ]
 
    data_tbl.setStyle(TableStyle(style_cmds))
 
    # ── Build PDF ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
    story = [
        summary_tbl,
        Spacer(1, 6 * mm),
        Paragraph("Dispatch Entry Statement", section_style),
        HRFlowable(width="100%", thickness=0.6, color=RULE, spaceAfter=4),
        data_tbl,
    ]
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export raw material report pdf.

def export_raw_material_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
    date_column_index: int = 0,
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY       = HexColor("#0D2545")
    NAVY_LIGHT = HexColor("#1A3A6B")
    GOLD       = HexColor("#C8922A")
    GOLD_LIGHT = HexColor("#F5E6C8")
    GREY_DARK  = HexColor("#3D3D3D")
    GREY_MID   = HexColor("#6B6B6B")
    GREY_LIGHT = HexColor("#F2F4F7")
    WHITE      = HexColor("#FFFFFF")
    ROW_ALT    = HexColor("#F8FAFC")
    RULE       = HexColor("#D9DDE6")
    DATE_BG    = HexColor("#E8EEF7")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # ── Group rows by date ────────────────────────────────────────────────────
    grouped: OrderedDict[str, list] = OrderedDict()
    for row in normalized_rows:
        date_key = str(row[date_column_index]).strip()
        try:
            display_date = datetime.strptime(date_key, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            display_date = date_key
        grouped.setdefault(display_date, []).append(row)
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        # Header band
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
 
        # Gold accent stripe
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        # Company name
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        # Report subtitle
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "RAW MATERIAL INWARD REPORT")
 
        # Meta — right side
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Entries: {len(normalized_rows)}")
 
        # Page number pill
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        # Footer band
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Raw Material Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_cell_style = ParagraphStyle(
        "RMHeaderCell", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    data_cell_style = ParagraphStyle(
        "RMDataCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    summary_label_style = ParagraphStyle(
        "RMSummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "RMSummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    date_heading_style = ParagraphStyle(
        "RMDateHeading", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", leading=13,
    )
    date_sub_style = ParagraphStyle(
        "RMDateSub", parent=styles["Normal"],
        fontSize=7.5, textColor=GREY_MID, leading=11,
    )
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("REPORT SUMMARY", summary_label_style),
        Paragraph(f"Total Entries: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Days: <b>{len(grouped)}</b>", summary_value_style),
        Paragraph(f"Generated: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.28, 0.22, 0.14, 0.36)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Helper: build date heading banner ─────────────────────────────────────

    def build_date_banner(display_date: str, entry_count: int) -> Table:
        banner_data = [[
            Paragraph(display_date, date_heading_style),
            Paragraph(f"{entry_count} entr{'y' if entry_count == 1 else 'ies'}", date_sub_style),
        ]]
        banner = Table(banner_data, colWidths=[AVAIL_W * 0.75, AVAIL_W * 0.25])
        banner.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), DATE_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, NAVY),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RULE),
            ("ALIGN",         (1, 0), (1, 0),   "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return banner
 
    # ── Assemble story ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
 
    story = [
        summary_tbl,
        Spacer(1, 5 * mm),
    ]
 
    for i, (display_date, day_rows) in enumerate(grouped.items()):
        banner = build_date_banner(display_date, len(day_rows))
        
        # Add banner
        story.append(banner)
        story.append(Spacer(1, 1.5 * mm))
        
        # Build combined table: header + all data rows
        col_w = AVAIL_W / len(headers)
        combined_data = [[Paragraph(str(h), header_cell_style) for h in headers]]
        
        for row in day_rows:
            combined_data.append([Paragraph(str(cell), data_cell_style) for cell in row])
        
        combined_tbl = Table(combined_data, colWidths=[col_w] * len(headers), repeatRows=1)
        combined_tbl.setStyle(TableStyle([
            # Header row styling
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1, 0),  8),
            ("TOPPADDING",     (0, 0), (-1, 0),  9),
            ("BOTTOMPADDING",  (0, 0), (-1, 0),  9),
            ("VALIGN",         (0, 0), (-1, 0),  "MIDDLE"),
            ("ALIGN",          (0, 0), (-1, 0),  "CENTER"),
            # Data rows
            ("TOPPADDING",     (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING",  (0, 1), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("TEXTCOLOR",      (0, 1), (-1, -1), GREY_DARK),
            ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",       (0, 1), (-1, -1), 8),
            ("VALIGN",         (0, 1), (-1, -1), "MIDDLE"),
            # Shared
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
            # Lines
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]))
        
        story.append(combined_tbl)
        
        # Gap between day sections (not after the last one)
        if i < len(grouped) - 1:
            story.append(Spacer(1, 6 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()
 
# sections: list of dicts, each with:
# {
# "title":   "Raw Material Entry Details",
# "headers": ["Field", "Value"],
# "rows":    [("Entry Code", "RMX00001"), ("Date", "2026-04-06"), ...],
# }
# Two expected sections:
# 1. Raw Material Entry Details  — Field / Value rows
# 2. Lab Report                  — Parameter / Value rows

def export_raw_material_entry_report_pdf(
    sections: list[dict],
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    FIELD_LABEL = HexColor("#1E3A5F")
    FIELD_BG    = HexColor("#EEF3FA")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    styles = getSampleStyleSheet()
 
    # Pull basic meta from first section for the header badge
    first_rows = list(sections[0].get("rows") or []) if sections else []
    meta = {str(r[0]): str(r[1]) for r in first_rows if len(r) >= 2}
    entry_code = meta.get("Entry Code") or "—"
    rm_type  = meta.get("RM Type",  "—")
    date_str = meta.get("Date",     "—")
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        # Header band
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "RAW MATERIAL ENTRY REPORT")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Entry Code: {entry_code}  |  {rm_type}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        # Footer
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Raw Material Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    label_style = ParagraphStyle(
        "RMELabel", parent=styles["Normal"],
        fontSize=8, textColor=FIELD_LABEL, fontName="Helvetica-Bold", leading=11,
    )
    value_style = ParagraphStyle(
        "RMEValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    section_title_style = ParagraphStyle(
        "RMESectionTitle", parent=styles["Normal"],
        fontSize=9, textColor=WHITE, fontName="Helvetica-Bold",
    )
 
    # ── Helper: section heading bar ───────────────────────────────────────────

    def build_section_heading(title: str) -> Table:
        t = Table([[Paragraph(title, section_title_style)]], colWidths=[AVAIL_W])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), NAVY),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  4, GOLD),
        ]))
        return t
 
    # ── Helper: Field/Value or Parameter/Value — 2-column card ───────────────

    def build_field_value_table(rows: list) -> Table:
        normalized = [list(r) for r in (rows or [])]
        if not normalized:
            normalized = [["No data available", "—"]]
 
        # Layout: pair rows into 4-column grid [label, value, label, value]
        paired = []
        for i in range(0, len(normalized), 2):
            left_k  = str(normalized[i][0])   if len(normalized[i]) > 0 else ""
            left_v  = str(normalized[i][1])   if len(normalized[i]) > 1 else ""
            if i + 1 < len(normalized):
                right_k = str(normalized[i+1][0]) if len(normalized[i+1]) > 0 else ""
                right_v = str(normalized[i+1][1]) if len(normalized[i+1]) > 1 else ""
            else:
                right_k, right_v = "", ""
 
            paired.append([
                Paragraph(left_k,  label_style),
                Paragraph(left_v,  value_style),
                Paragraph(right_k, label_style),
                Paragraph(right_v, value_style),
            ])
 
        col_widths = [AVAIL_W * w for w in (0.22, 0.28, 0.22, 0.28)]
        tbl = Table(paired, colWidths=col_widths)
 
        row_cmds = [
            ("BACKGROUND", (0, i), (-1, i), FIELD_BG if i % 2 == 0 else WHITE)
            for i in range(len(paired))
        ]
        tbl.setStyle(TableStyle([
            # Label columns always tinted
            ("BACKGROUND",    (0, 0), (0, -1),  FIELD_BG),
            ("BACKGROUND",    (2, 0), (2, -1),  FIELD_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 9),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 9),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("LINEAFTER",     (0, 0), (0, -1),  0.4, RULE),
            ("LINEAFTER",     (1, 0), (1, -1),  1.0, RULE),   # centre divider
            ("LINEAFTER",     (2, 0), (2, -1),  0.4, RULE),
            ("LINEBELOW",     (0, 0), (-1, -2), 0.4, RULE),
            ("BOX",           (0, 0), (-1, -1), 0.8, RULE),
        ] + row_cmds))
        return tbl
 
    # ── Entry banner ──────────────────────────────────────────────────────────
    banner_id_style = ParagraphStyle(
        "RMEBannerId", parent=styles["Normal"],
        fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", leading=15,
    )
    banner_sub_style = ParagraphStyle(
        "RMEBannerSub", parent=styles["Normal"],
        fontSize=7.5, textColor=GREY_MID, leading=11,
    )
    banner_tag_style = ParagraphStyle(
        "RMEBannerTag", parent=styles["Normal"],
        fontSize=7, textColor=NAVY, fontName="Helvetica-Bold", alignment=2,
    )
 
    banner_data = [[
        [
            Paragraph(f"Entry #{entry_code} — {rm_type}", banner_id_style),
            Paragraph(f"Inward Date: {date_str}", banner_sub_style),
        ],
        [Paragraph("ENTRY RECORD", banner_tag_style)],
    ]]
    banner = Table(banner_data, colWidths=[AVAIL_W * 0.72, AVAIL_W * 0.28])
    banner.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LINEBEFORE",    (0, 0), (0, -1),  4, GOLD),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (1, 0), (1, -1),  "RIGHT"),
    ]))
 
    # ── Assemble story ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
 
    story = [banner, Spacer(1, 5 * mm)]
 
    for i, section in enumerate(sections or []):
        title   = section.get("title") or "Section"
        rows    = section.get("rows") or []
 
        story.append(KeepTogether([
            build_section_heading(title),
            Spacer(1, 1.5 * mm),
            build_field_value_table(rows),
        ]))
 
        if i < len(sections) - 1:
            story.append(Spacer(1, 5 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export production report pdf.

def export_production_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
    date_column_index: int = 0,
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    DATE_BG     = HexColor("#E8EEF7")
 
    # ── Constants — landscape A4 ──────────────────────────────────────────────
    PAGE_SIZE  = A4                          # ← remove landscape()
    PAGE_W, PAGE_H = PAGE_SIZE
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # ── Group rows by date ────────────────────────────────────────────────────
    grouped: OrderedDict[str, list] = OrderedDict()
    for row in normalized_rows:
        date_key = str(row[date_column_index]).strip()
        try:
            display_date = datetime.strptime(date_key, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            display_date = date_key
        grouped.setdefault(display_date, []).append(row)
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = PAGE_SIZE
 
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "PRODUCTION REPORT")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Batches: {len(normalized_rows)}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Production Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_cell_style  = ParagraphStyle("PRHeaderCell", parent=styles["Normal"], fontSize=7,   textColor=WHITE,     fontName="Helvetica-Bold", leading=9,  alignment=1)
    data_cell_style    = ParagraphStyle("PRDataCell",   parent=styles["Normal"], fontSize=7,   textColor=GREY_DARK, leading=9)
    summary_label_style= ParagraphStyle("PRSummaryLabel",parent=styles["Normal"],fontSize=8,   textColor=NAVY,      fontName="Helvetica-Bold")
    summary_value_style= ParagraphStyle("PRSummaryValue",parent=styles["Normal"],fontSize=8,   textColor=GREY_MID)
    date_heading_style = ParagraphStyle("PRDateHeading", parent=styles["Normal"],fontSize=9,   textColor=NAVY,      fontName="Helvetica-Bold", leading=13)
    date_sub_style     = ParagraphStyle("PRDateSub",     parent=styles["Normal"],fontSize=7.5, textColor=GREY_MID,  leading=11)
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("PRODUCTION SUMMARY", summary_label_style),
        Paragraph(f"Total Batches: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Days: <b>{len(grouped)}</b>", summary_value_style),
        Paragraph(f"Generated: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(summary_data, colWidths=[AVAIL_W * w for w in (0.28, 0.20, 0.14, 0.38)])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),  ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Helper: smart column widths ───────────────────────────────────────────

    def compute_col_widths(n_cols):
        if n_cols == 9:
            # Date, Batch No, Product, Batch Size, MOP, Water, Bags, Wt/Bag, Output
            return [w * AVAIL_W for w in (0.13, 0.11, 0.18, 0.10, 0.09, 0.09, 0.10, 0.10, 0.10)]
        return [AVAIL_W / n_cols] * n_cols
    
    # ── Helper: day data table ────────────────────────────────────────────────

    def build_day_table(day_rows):
        col_widths = compute_col_widths(max(len(headers), 1))
        td = [[Paragraph(str(h), header_cell_style) for h in headers]]
        for row in day_rows:
            td.append([Paragraph(str(cell), data_cell_style) for cell in row])
 
        tbl = Table(td, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TOPPADDING",     (0, 0), (-1, 0),  7), ("BOTTOMPADDING", (0, 0), (-1, 0),  7),
            ("TOPPADDING",     (0, 1), (-1, -1), 5), ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 5), ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]))
        return tbl
 
    # ── Helper: date banner ───────────────────────────────────────────────────

    def build_date_banner(display_date, row_count):
        bd = [[
            Paragraph(display_date, date_heading_style),
            Paragraph(f"{row_count} batch{'es' if row_count != 1 else ''}", date_sub_style),
        ]]
        b = Table(bd, colWidths=[AVAIL_W * 0.75, AVAIL_W * 0.25])
        b.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), DATE_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),  ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, NAVY),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RULE),
            ("ALIGN",         (1, 0), (1, 0),   "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return b
 
    # ── Assemble story ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=PAGE_SIZE,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
 
    story = [summary_tbl, Spacer(1, 7 * mm)]
    for i, (display_date, day_rows) in enumerate(grouped.items()):
        story.append(KeepTogether([
            build_date_banner(display_date, len(day_rows)),
            Spacer(1, 1.5 * mm),
            build_day_table(day_rows),
        ]))
        if i < len(grouped) - 1:
            story.append(Spacer(1, 6 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# sections: list of dicts, each with:
# {
# "title":   "Batch Details" | "Consumption Details" | etc.,
# "headers": [...],
# "rows":    [...],
# }
# Expected:
# sections[0] = Batch Details       → rendered as card grid (like image 2)
# sections[1] = Consumption Details → rendered as table with TOTAL row

def export_batch_consumption_report_pdf(
    sections: list[dict],
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    CARD_BG     = HexColor("#F0F4FA")
    CARD_LABEL  = HexColor("#1E3A5F")
    TOTAL_BG    = HexColor("#EEF3FA")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    styles       = getSampleStyleSheet()
 
    # Pull meta from first section for header badge
    first_rows = list(sections[0].get("rows") or []) if sections else []
    meta = {str(r[0]): str(r[1]) for r in first_rows if len(r) >= 2}
    batch_no  = meta.get("Batch No",  meta.get("Batch No", "—"))
    product   = meta.get("Product",   "—")
    date_str  = meta.get("Date",      "—")
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "BATCH CONSUMPTION REPORT")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Batch: {batch_no}  |  {product}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Production Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    card_label_style = ParagraphStyle(
        "BCCardLabel", parent=styles["Normal"],
        fontSize=7, textColor=GREY_MID, fontName="Helvetica", leading=10,
    )
    card_value_style = ParagraphStyle(
        "BCCardValue", parent=styles["Normal"],
        fontSize=9, textColor=CARD_LABEL, fontName="Helvetica-Bold", leading=13,
    )
    section_heading_style = ParagraphStyle(
        "BCSectionHeading", parent=styles["Normal"],
        fontSize=9, textColor=WHITE, fontName="Helvetica-Bold",
    )
    header_cell_style = ParagraphStyle(
        "BCHeaderCell", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    data_cell_style = ParagraphStyle(
        "BCDataCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    total_style = ParagraphStyle(
        "BCTotal", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold", leading=11,
    )
 
    # ── Helper: section heading bar ───────────────────────────────────────────

    def build_section_heading(title: str) -> Table:
        t = Table([[Paragraph(title, section_heading_style)]], colWidths=[AVAIL_W])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), NAVY),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  4, GOLD),
        ]))
        return t
 
    # ── Helper: card grid (Image 2 style) ─────────────────────────────────────
    # Lays out field/value pairs in rows of `cols_per_row` cards side by side

    def build_card_grid(rows: list, cols_per_row: int = 5) -> Table:
        items = [(str(r[0]), str(r[1])) for r in (rows or []) if len(r) >= 2]
        if not items:
            items = [("No data", "—")]
 
        # Pad to fill last row
        while len(items) % cols_per_row != 0:
            items.append(("", ""))
 
        card_w = AVAIL_W / cols_per_row
 
        grid_rows = []
        for i in range(0, len(items), cols_per_row):
            chunk = items[i:i + cols_per_row]
            # Label row
            grid_rows.append([Paragraph(label, card_label_style) for label, _ in chunk])
            # Value row
            grid_rows.append([Paragraph(value, card_value_style) for _, value in chunk])
 
        tbl = Table(grid_rows, colWidths=[card_w] * cols_per_row)
 
        style_cmds = [
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("BOX",           (0, 0), (-1, -1), 0.8, RULE),
            ("LINEAFTER",     (0, 0), (-2, -1), 0.5, RULE),
        ]
        # Each pair of rows (label+value) = one card row — add bottom border between card rows
        for r in range(1, len(grid_rows) - 1, 2):
            style_cmds.append(("LINEBELOW", (0, r), (-1, r), 0.6, RULE))
        # Extra top padding on label rows, extra bottom padding on value rows
        for r in range(0, len(grid_rows), 2):   # label rows
            style_cmds.append(("TOPPADDING",    (0, r), (-1, r), 8))
            style_cmds.append(("BOTTOMPADDING", (0, r), (-1, r), 1))
        for r in range(1, len(grid_rows), 2):   # value rows
            style_cmds.append(("TOPPADDING",    (0, r), (-1, r), 1))
            style_cmds.append(("BOTTOMPADDING", (0, r), (-1, r), 8))
            style_cmds.append(("BACKGROUND",    (0, r-1), (-1, r), CARD_BG))
 
        tbl.setStyle(TableStyle(style_cmds))
        return tbl
 
    # ── Helper: consumption / regular table with optional TOTAL row ───────────

    def build_data_table(rows: list, headers: list) -> Table:
        normalized = [list(r) for r in (rows or [])]
        if not normalized:
            normalized = [["No data available"]]
 
        # Detect TOTAL row
        data_rows  = normalized
        total_row  = None
        if normalized and str(normalized[-1][0]).strip().upper() == "TOTAL":
            data_rows = normalized[:-1]
            total_row = normalized[-1]
 
        col_w = AVAIL_W / max(len(headers), 1)
        td = [[Paragraph(str(h), header_cell_style) for h in headers]]
        for row in data_rows:
            td.append([Paragraph(str(cell), data_cell_style) for cell in row])
 
        total_idx = None
        if total_row is not None:
            total_idx = len(td)
            td.append([Paragraph(str(cell), total_style) for cell in total_row])
 
        tbl = Table(td, colWidths=[col_w] * len(headers), repeatRows=1)
 
        style_cmds = [
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TOPPADDING",     (0, 0), (-1, 0),  9), ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
            ("TOPPADDING",     (0, 1), (-1, -1), 6), ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8), ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]
        if total_idx is not None:
            style_cmds += [
                ("BACKGROUND", (0, total_idx), (-1, total_idx), TOTAL_BG),
                ("LINEABOVE",  (0, total_idx), (-1, total_idx), 1.2, NAVY),
                ("LINEBELOW",  (0, total_idx), (-1, total_idx), 1.2, NAVY),
            ]
        tbl.setStyle(TableStyle(style_cmds))
        return tbl
 
    # ── Assemble story ────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
 
    story = []
 
    for i, section in enumerate(sections or []):
        title   = section.get("title") or "Section"
        headers = section.get("headers") or []
        rows    = section.get("rows") or []
 
        is_details = i == 0  # first section → card grid
 
        if is_details:
            block = [
                build_section_heading(title),
                Spacer(1, 2 * mm),
                build_card_grid(rows, cols_per_row=5),
            ]
        else:
            block = [
                build_section_heading(title),
                Spacer(1, 2 * mm),
                build_data_table(rows, headers),
            ]
 
        story.append(KeepTogether(block))
 
        if i < len(sections) - 1:
            story.append(Spacer(1, 5 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export rm stock report pdf.

def export_rm_stock_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
    date_column_index: int = 0,
) -> bytes:
 
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    DATE_BG     = HexColor("#E8EEF7")
 
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    grouped: OrderedDict[str, list] = OrderedDict()
    for row in normalized_rows:
        date_key = str(row[date_column_index]).strip()
        try:
            display_date = datetime.strptime(date_key, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            display_date = date_key
        grouped.setdefault(display_date, []).append(row)
 
    # Handle decorator.

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "RAW MATERIAL STOCK REPORT")
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Rows: {len(normalized_rows)}")
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm, f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Raw Material Management System")
        canvas.restoreState()
 
    header_cell_style = ParagraphStyle("RMSRHeaderCell", parent=styles["Normal"], fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11)
    data_cell_style   = ParagraphStyle("RMSRDataCell",   parent=styles["Normal"], fontSize=8, textColor=GREY_DARK, leading=11)
    summary_label_style = ParagraphStyle("RMSRSummaryLabel", parent=styles["Normal"], fontSize=8, textColor=NAVY, fontName="Helvetica-Bold")
    summary_value_style = ParagraphStyle("RMSRSummaryValue", parent=styles["Normal"], fontSize=8, textColor=GREY_MID)
    date_heading_style  = ParagraphStyle("RMSRDateHeading",  parent=styles["Normal"], fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", leading=13)
    date_sub_style      = ParagraphStyle("RMSRDateSub",      parent=styles["Normal"], fontSize=7.5, textColor=GREY_MID, leading=11)
 
    summary_data = [[
        Paragraph("RM STOCK LEDGER", summary_label_style),
        Paragraph(f"Total Rows: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Days: <b>{len(grouped)}</b>", summary_value_style),
        Paragraph(f"Generated: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(summary_data, colWidths=[AVAIL_W * w for w in (0.28, 0.18, 0.18, 0.36)])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),  ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # Build day table.

    def build_day_table(day_rows):
        col_w = AVAIL_W / max(len(headers), 1)
        td = [[Paragraph(str(h), header_cell_style) for h in headers]]
        for row in day_rows:
            td.append([Paragraph(str(cell), data_cell_style) for cell in row])
        tbl = Table(td, colWidths=[col_w] * len(headers), repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TOPPADDING",     (0, 0), (-1, 0),  9), ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
            ("TOPPADDING",     (0, 1), (-1, -1), 6), ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]))
        return tbl
 
    # Build date banner.

    def build_date_banner(display_date, row_count):
        bd = [[
            Paragraph(display_date, date_heading_style),
            Paragraph(f"{row_count} entr{'y' if row_count == 1 else 'ies'}", date_sub_style),
        ]]
        b = Table(bd, colWidths=[AVAIL_W * 0.75, AVAIL_W * 0.25])
        b.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), DATE_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),  ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, NAVY),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RULE),
            ("ALIGN",         (1, 0), (1, 0),   "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return b
 
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=MARGIN_H, rightMargin=MARGIN_H, topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT)
 
    story = [summary_tbl, Spacer(1, 7 * mm)]
    for i, (display_date, day_rows) in enumerate(grouped.items()):
        story.append(KeepTogether([
            build_date_banner(display_date, len(day_rows)),
            Spacer(1, 1.5 * mm),
            build_day_table(day_rows),
        ]))
        if i < len(grouped) - 1:
            story.append(Spacer(1, 6 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export rm individual stock report pdf.

def export_rm_individual_stock_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    ZERO_FG     = HexColor("#991B1B")   # red text for zero/empty stock
    ZERO_BG     = HexColor("#FEF2F2")
    LOW_FG      = HexColor("#92400E")   # amber for low stock
    LOW_BG      = HexColor("#FFFBEB")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "RAW MATERIAL STOCK SUMMARY")
 
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Items: {len(normalized_rows)}")
 
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Raw Material Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_style = ParagraphStyle(
        "RMSHeader", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    cell_style = ParagraphStyle(
        "RMSCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    zero_style = ParagraphStyle(
        "RMSZero", parent=styles["Normal"],
        fontSize=8, textColor=ZERO_FG, fontName="Helvetica-Bold", leading=11,
    )
    low_style = ParagraphStyle(
        "RMSLow", parent=styles["Normal"],
        fontSize=8, textColor=LOW_FG, fontName="Helvetica-Bold", leading=11,
    )
    summary_label_style = ParagraphStyle(
        "RMSSummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "RMSSummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    section_style = ParagraphStyle(
        "RMSSection", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4,
    )
 
    # ── Compute summary stats ─────────────────────────────────────────────────
    stock_values = []
    for r in normalized_rows:
        try:
            stock_values.append(float(r[1]))
        except (IndexError, ValueError, TypeError):
            stock_values.append(0.0)
 
    total_stock  = sum(stock_values)
    zero_count   = sum(1 for v in stock_values if v == 0)
    low_count    = sum(1 for v in stock_values if 0 < v < 500)
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("STOCK OVERVIEW", summary_label_style),
        Paragraph(f"Total Items: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Total Stock: <b>{total_stock:,.1f} kg</b>", summary_value_style),
        Paragraph(f"Zero Stock: <b>{zero_count}</b>", summary_value_style),
        Paragraph(f"Low Stock (&lt;500kg): <b>{low_count}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.22, 0.18, 0.24, 0.18, 0.18)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Stock table ───────────────────────────────────────────────────────────
    # Use a wider left column for RM name, narrower right for stock value
    COL_NAME  = AVAIL_W * 0.60
    COL_STOCK = AVAIL_W * 0.40
 
    table_data = [[Paragraph(str(h), header_style) for h in (headers or ["RM Type", "Current Stock (kg)"])]]
 
    row_cmds = []
    for i, (row, stock_val) in enumerate(zip(normalized_rows, stock_values), start=1):
        name = str(row[0]) if len(row) > 0 else "—"
        val  = str(row[1]) if len(row) > 1 else "—"
 
        if stock_val == 0:
            name_p = Paragraph(name, zero_style)
            val_p  = Paragraph(val,  zero_style)
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), ZERO_BG))
        elif stock_val < 500:
            name_p = Paragraph(name, low_style)
            val_p  = Paragraph(val,  low_style)
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), LOW_BG))
        else:
            name_p = Paragraph(name, cell_style)
            val_p  = Paragraph(val,  cell_style)
            bg = WHITE if i % 2 == 1 else ROW_ALT
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
 
        table_data.append([name_p, val_p])
 
    stock_tbl = Table(table_data, colWidths=[COL_NAME, COL_STOCK], repeatRows=1)
    stock_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
        ("TOPPADDING",    (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
        ("TOPPADDING",    (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (1, 0), (1, -1),  "RIGHT"),   # stock values right-aligned
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, GOLD),
        ("LINEBELOW",     (0, 1), (-1, -1), 0.4, RULE),
        ("LINEAFTER",     (0, 0), (0, -1),  0.4, RULE),
        ("BOX",           (0, 0), (-1, -1), 0.8, RULE),
    ] + row_cmds))
 
    # ── Legend ────────────────────────────────────────────────────────────────
    legend_data = [[
        Paragraph("■", ParagraphStyle("lz", parent=styles["Normal"], fontSize=8, textColor=ZERO_FG)),
        Paragraph("Zero Stock", ParagraphStyle("lzt", parent=styles["Normal"], fontSize=7.5, textColor=ZERO_FG)),
        Paragraph("■", ParagraphStyle("ll", parent=styles["Normal"], fontSize=8, textColor=LOW_FG)),
        Paragraph("Low Stock (&lt;500 kg)", ParagraphStyle("llt", parent=styles["Normal"], fontSize=7.5, textColor=LOW_FG)),
    ]]
    legend = Table(legend_data, colWidths=[5 * mm, 30 * mm, 5 * mm, 40 * mm])
    legend.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Build PDF ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
    story = [
        summary_tbl,
        Spacer(1, 6 * mm),
        Paragraph("Current Stock Levels", section_style),
        HRFlowable(width="100%", thickness=0.6, color=RULE, spaceAfter=4),
        stock_tbl,
        Spacer(1, 3 * mm),
        legend,
    ]
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export feed individual stock report pdf.

def export_feed_individual_stock_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    ZERO_FG     = HexColor("#991B1B")   # red text for zero/empty stock
    ZERO_BG     = HexColor("#FEF2F2")
    LOW_FG      = HexColor("#92400E")   # amber for low stock
    LOW_BG      = HexColor("#FFFBEB")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        # Header bar
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        # Company name
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        # Report subtitle
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "FEED STOCK SUMMARY")
 
        # Meta info (right-aligned)
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Items: {len(normalized_rows)}")
 
        # Page number badge
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        # Footer bar
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Feed Stock Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_style = ParagraphStyle(
        "FeedHeader", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    cell_style = ParagraphStyle(
        "FeedCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    zero_style = ParagraphStyle(
        "FeedZero", parent=styles["Normal"],
        fontSize=8, textColor=ZERO_FG, fontName="Helvetica-Bold", leading=11,
    )
    low_style = ParagraphStyle(
        "FeedLow", parent=styles["Normal"],
        fontSize=8, textColor=LOW_FG, fontName="Helvetica-Bold", leading=11,
    )
    summary_label_style = ParagraphStyle(
        "FeedSummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "FeedSummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    section_style = ParagraphStyle(
        "FeedSection", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4,
    )
 
    # ── Compute summary stats ─────────────────────────────────────────────────
    # Stock value is the LAST column (index -1 or index 2 for a 3-col layout)
    stock_values = []
    for r in normalized_rows:
        try:
            stock_values.append(float(r[-1]))
        except (IndexError, ValueError, TypeError):
            stock_values.append(0.0)
 
    total_stock = sum(stock_values)
    zero_count  = sum(1 for v in stock_values if v == 0)
    low_count   = sum(1 for v in stock_values if 0 < v < 500)
 
    # ── Summary strip ─────────────────────────────────────────────────────────
    summary_data = [[
        Paragraph("STOCK OVERVIEW", summary_label_style),
        Paragraph(f"Total Items: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Total Stock: <b>{total_stock:,.1f} kg</b>", summary_value_style),
        Paragraph(f"Zero Stock: <b>{zero_count}</b>", summary_value_style),
        Paragraph(f"Low Stock (&lt;500kg): <b>{low_count}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.22, 0.18, 0.24, 0.18, 0.18)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Feed table ────────────────────────────────────────────────────────────
    # 3-column layout: Feed Type | Bag Size Mix | Available Stock (kg)
    resolved_headers = headers or ["Feed Type", "Bag Size Mix", "Available Stock (kg)"]
    num_cols = len(resolved_headers)
 
    if num_cols == 3:
        col_widths = [AVAIL_W * 0.42, AVAIL_W * 0.28, AVAIL_W * 0.30]
    elif num_cols == 2:
        col_widths = [AVAIL_W * 0.60, AVAIL_W * 0.40]
    else:
        # Distribute evenly
        col_widths = [AVAIL_W / num_cols] * num_cols
 
    table_data = [[Paragraph(str(h), header_style) for h in resolved_headers]]
 
    row_cmds = []
    for i, (row, stock_val) in enumerate(zip(normalized_rows, stock_values), start=1):
        # Pad or trim row to match column count
        padded = list(row) + ["—"] * num_cols
        padded = padded[:num_cols]
 
        if stock_val == 0:
            styled = [Paragraph(str(v), zero_style) for v in padded]
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), ZERO_BG))
        elif stock_val < 500:
            styled = [Paragraph(str(v), low_style) for v in padded]
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), LOW_BG))
        else:
            styled = [Paragraph(str(v), cell_style) for v in padded]
            bg = WHITE if i % 2 == 1 else ROW_ALT
            row_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
 
        table_data.append(styled)
 
    feed_tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
 
    # Build vertical divider commands for every interior column boundary
    divider_cmds = [
        ("LINEAFTER", (c, 0), (c, -1), 0.4, RULE)
        for c in range(num_cols - 1)
    ]
 
    feed_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
        ("TOPPADDING",    (0, 0), (-1, 0),  9),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
        ("TOPPADDING",    (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN",         (num_cols - 1, 0), (num_cols - 1, -1), "RIGHT"),   # stock col right-aligned
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, GOLD),
        ("LINEBELOW",     (0, 1), (-1, -1), 0.4, RULE),
        ("BOX",           (0, 0), (-1, -1), 0.8, RULE),
    ] + divider_cmds + row_cmds))
 
    # ── Legend ────────────────────────────────────────────────────────────────
    legend_data = [[
        Paragraph("■", ParagraphStyle("flz",  parent=styles["Normal"], fontSize=8,   textColor=ZERO_FG)),
        Paragraph("Zero Stock",         ParagraphStyle("flzt", parent=styles["Normal"], fontSize=7.5, textColor=ZERO_FG)),
        Paragraph("■", ParagraphStyle("fll",  parent=styles["Normal"], fontSize=8,   textColor=LOW_FG)),
        Paragraph("Low Stock (&lt;500 kg)", ParagraphStyle("fllt", parent=styles["Normal"], fontSize=7.5, textColor=LOW_FG)),
    ]]
    legend = Table(legend_data, colWidths=[5 * mm, 30 * mm, 5 * mm, 40 * mm])
    legend.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Build PDF ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
    story = [
        summary_tbl,
        Spacer(1, 6 * mm),
        Paragraph("Current Feed Stock Levels", section_style),
        HRFlowable(width="100%", thickness=0.6, color=RULE, spaceAfter=4),
        feed_tbl,
        Spacer(1, 3 * mm),
        legend,
    ]
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export feed stock report pdf.

def export_feed_stock_report_pdf(
    headers: list,
    rows: list,
    company_name: str = "POULTRY NET",
    date_column_index: int = 0,
) -> bytes:
 
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    DATE_BG     = HexColor("#E8EEF7")
 
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at    = datetime.now().strftime("%d %B %Y, %H:%M")
    normalized_rows = [list(r) if not isinstance(r, list) else r for r in rows]
    styles          = getSampleStyleSheet()
 
    grouped: OrderedDict[str, list] = OrderedDict()
    for row in normalized_rows:
        date_key = str(row[date_column_index]).strip()
        try:
            display_date = datetime.strptime(date_key, "%Y-%m-%d").strftime("%d %B %Y")
        except ValueError:
            display_date = date_key
        grouped.setdefault(display_date, []).append(row)
 
    # Handle decorator.

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "FEED STOCK REPORT")
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm, f"Total Rows: {len(normalized_rows)}")
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm, f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Feed Management System")
        canvas.restoreState()
 
    header_cell_style  = ParagraphStyle("FSHeaderCell",    parent=styles["Normal"], fontSize=8, textColor=WHITE,     fontName="Helvetica-Bold", leading=11)
    data_cell_style    = ParagraphStyle("FSDataCell",      parent=styles["Normal"], fontSize=8, textColor=GREY_DARK, leading=11)
    summary_label_style= ParagraphStyle("FSSummaryLabel",  parent=styles["Normal"], fontSize=8, textColor=NAVY,      fontName="Helvetica-Bold")
    summary_value_style= ParagraphStyle("FSSummaryValue",  parent=styles["Normal"], fontSize=8, textColor=GREY_MID)
    date_heading_style = ParagraphStyle("FSDateHeading",   parent=styles["Normal"], fontSize=9, textColor=NAVY,      fontName="Helvetica-Bold", leading=13)
    date_sub_style     = ParagraphStyle("FSDateSub",       parent=styles["Normal"], fontSize=7.5, textColor=GREY_MID, leading=11)
 
    summary_data = [[
        Paragraph("FEED STOCK SUMMARY", summary_label_style),
        Paragraph(f"Total Rows: <b>{len(normalized_rows)}</b>", summary_value_style),
        Paragraph(f"Days: <b>{len(grouped)}</b>", summary_value_style),
        Paragraph(f"Generated: <b>{generated_at}</b>", summary_value_style),
    ]]
    summary_tbl = Table(summary_data, colWidths=[AVAIL_W * w for w in (0.28, 0.18, 0.18, 0.36)])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),  ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # Build day table.

    def build_day_table(day_rows):
        col_w = AVAIL_W / max(len(headers), 1)
        td = [[Paragraph(str(h), header_cell_style) for h in headers]]
        for row in day_rows:
            td.append([Paragraph(str(cell), data_cell_style) for cell in row])
        tbl = Table(td, colWidths=[col_w] * len(headers), repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  NAVY),
            ("TOPPADDING",     (0, 0), (-1, 0),  9), ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
            ("TOPPADDING",     (0, 1), (-1, -1), 6), ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, ROW_ALT]),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8), ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",      (0, 1), (-1, -1), 0.4, RULE),
            ("LINEAFTER",      (0, 0), (-2, -1), 0.4, RULE),
            ("BOX",            (0, 0), (-1, -1), 0.8, RULE),
        ]))
        return tbl
 
    # Build date banner.

    def build_date_banner(display_date, row_count):
        bd = [[
            Paragraph(display_date, date_heading_style),
            Paragraph(f"{row_count} entr{'y' if row_count == 1 else 'ies'}", date_sub_style),
        ]]
        b = Table(bd, colWidths=[AVAIL_W * 0.75, AVAIL_W * 0.25])
        b.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), DATE_BG),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 7),  ("BOTTOMPADDING",(0, 0), (-1, -1), 7),
            ("LINEBEFORE",    (0, 0), (0, -1),  3, NAVY),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.5, RULE),
            ("ALIGN",         (1, 0), (1, 0),   "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return b
 
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=MARGIN_H, rightMargin=MARGIN_H, topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT)
 
    story = [summary_tbl, Spacer(1, 7 * mm)]
    for i, (display_date, day_rows) in enumerate(grouped.items()):
        story.append(KeepTogether([
            build_date_banner(display_date, len(day_rows)),
            Spacer(1, 1.5 * mm),
            build_day_table(day_rows),
        ]))
        if i < len(grouped) - 1:
            story.append(Spacer(1, 6 * mm))
 
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export overall stock report pdf.

def export_overall_stock_report_pdf(
    sections: list[dict],
    company_name: str = "POULTRY NET",
) -> bytes:
 
    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY        = HexColor("#0D2545")
    NAVY_LIGHT  = HexColor("#1A3A6B")
    GOLD        = HexColor("#C8922A")
    GOLD_LIGHT  = HexColor("#F5E6C8")
    GREY_DARK   = HexColor("#3D3D3D")
    GREY_MID    = HexColor("#6B6B6B")
    GREY_LIGHT  = HexColor("#F2F4F7")
    WHITE       = HexColor("#FFFFFF")
    ROW_ALT     = HexColor("#F8FAFC")
    RULE        = HexColor("#D9DDE6")
    ZERO_FG     = HexColor("#991B1B")   # red text for zero/empty stock
    ZERO_BG     = HexColor("#FEF2F2")
    LOW_FG      = HexColor("#92400E")   # amber for low stock
    LOW_BG      = HexColor("#FFFBEB")
 
    # ── Constants ─────────────────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4
    MARGIN_H   = 18 * mm
    MARGIN_TOP = 36 * mm
    MARGIN_BOT = 20 * mm
    AVAIL_W    = PAGE_W - 2 * MARGIN_H
 
    generated_at = datetime.now().strftime("%d %B %Y, %H:%M")
    styles       = getSampleStyleSheet()
 
    # ── Normalise all sections up-front (needed for summary stats + decorator) ─
    normalised_sections = []
    for sec in (sections or []):
        title   = sec.get("title") or "Section"
        headers = list(sec.get("headers") or [])
        raw     = sec.get("rows") or []
        rows    = [list(r) if not isinstance(r, list) else r for r in raw]
        if not rows:
            rows    = [["No data available"]]
            headers = headers or ["Data"]
        normalised_sections.append({"title": title, "headers": headers, "rows": rows})
 
    # ── Aggregate counts across all sections for the summary strip ────────────
    all_stock_values = []
    total_item_count = 0
    for sec in normalised_sections:
        for row in sec["rows"]:
            total_item_count += 1
            try:
                all_stock_values.append(float(row[-1]))
            except (IndexError, ValueError, TypeError):
                all_stock_values.append(0.0)
 
    total_stock = sum(all_stock_values)
    zero_count  = sum(1 for v in all_stock_values if v == 0)
    low_count   = sum(1 for v in all_stock_values if 0 < v < 500)
 
    # ── Page decorator ────────────────────────────────────────────────────────

    def decorator(canvas, doc):
        canvas.saveState()
        W, H = A4
 
        # Header bar
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
 
        # Company name
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
 
        # Report subtitle
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "OVERALL STOCK REPORT")
 
        # Meta info (right-aligned)
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.drawRightString(meta_x, H - 18 * mm,
            f"Sections: {len(normalised_sections)}  |  Total Items: {total_item_count}")
 
        # Page number badge
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
 
        # Footer bar
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Stock Management System")
 
        canvas.restoreState()
 
    # ── Paragraph styles ──────────────────────────────────────────────────────
    header_style = ParagraphStyle(
        "OvrHeader", parent=styles["Normal"],
        fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", leading=11,
    )
    cell_style = ParagraphStyle(
        "OvrCell", parent=styles["Normal"],
        fontSize=8, textColor=GREY_DARK, leading=11,
    )
    zero_style = ParagraphStyle(
        "OvrZero", parent=styles["Normal"],
        fontSize=8, textColor=ZERO_FG, fontName="Helvetica-Bold", leading=11,
    )
    low_style = ParagraphStyle(
        "OvrLow", parent=styles["Normal"],
        fontSize=8, textColor=LOW_FG, fontName="Helvetica-Bold", leading=11,
    )
    summary_label_style = ParagraphStyle(
        "OvrSummaryLabel", parent=styles["Normal"],
        fontSize=8, textColor=NAVY, fontName="Helvetica-Bold",
    )
    summary_value_style = ParagraphStyle(
        "OvrSummaryValue", parent=styles["Normal"],
        fontSize=8, textColor=GREY_MID,
    )
    section_title_style = ParagraphStyle(
        "OvrSectionTitle", parent=styles["Normal"],
        fontSize=9, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4,
    )
 
    # ── Summary strip (combined totals across all sections) ───────────────────
    summary_data = [[
        Paragraph("STOCK OVERVIEW", summary_label_style),
        Paragraph(f"Sections: <b>{len(normalised_sections)}</b>", summary_value_style),
        Paragraph(f"Total Stock: <b>{total_stock:,.1f} kg</b>", summary_value_style),
        Paragraph(f"Zero Stock: <b>{zero_count}</b>", summary_value_style),
        Paragraph(f"Low Stock (&lt;500kg): <b>{low_count}</b>", summary_value_style),
    ]]
    summary_tbl = Table(
        summary_data,
        colWidths=[AVAIL_W * w for w in (0.22, 0.18, 0.24, 0.18, 0.18)],
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Helper: build one section's styled table ───────────────────────────────

    def build_section_table(headers, rows):
        num_cols   = len(headers)
        col_widths = [AVAIL_W / num_cols] * num_cols  # distribute evenly
 
        # Stock col is assumed to be the last column
        table_data = [[Paragraph(str(h), header_style) for h in headers]]
        row_cmds   = []
 
        for i, row in enumerate(rows, start=1):
            padded = (list(row) + ["—"] * num_cols)[:num_cols]
 
            try:
                stock_val = float(row[-1])
            except (IndexError, ValueError, TypeError):
                stock_val = 0.0
 
            if stock_val == 0:
                styled = [Paragraph(str(v), zero_style) for v in padded]
                row_cmds.append(("BACKGROUND", (0, i), (-1, i), ZERO_BG))
            elif stock_val < 500:
                styled = [Paragraph(str(v), low_style) for v in padded]
                row_cmds.append(("BACKGROUND", (0, i), (-1, i), LOW_BG))
            else:
                styled = [Paragraph(str(v), cell_style) for v in padded]
                bg = WHITE if i % 2 == 1 else ROW_ALT
                row_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))
 
            table_data.append(styled)
 
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
 
        divider_cmds = [
            ("LINEAFTER", (c, 0), (c, -1), 0.4, RULE)
            for c in range(num_cols - 1)
        ]
 
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("TOPPADDING",    (0, 0), (-1, 0),  9),
            ("BOTTOMPADDING", (0, 0), (-1, 0),  9),
            ("TOPPADDING",    (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",         (num_cols - 1, 0), (num_cols - 1, -1), "RIGHT"),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.5, GOLD),
            ("LINEBELOW",     (0, 1), (-1, -1), 0.4, RULE),
            ("BOX",           (0, 0), (-1, -1), 0.8, RULE),
        ] + divider_cmds + row_cmds))
 
        return tbl
 
    # ── Legend ────────────────────────────────────────────────────────────────
    legend_data = [[
        Paragraph("■", ParagraphStyle("olz",  parent=styles["Normal"], fontSize=8,   textColor=ZERO_FG)),
        Paragraph("Zero Stock",             ParagraphStyle("olzt", parent=styles["Normal"], fontSize=7.5, textColor=ZERO_FG)),
        Paragraph("■", ParagraphStyle("oll",  parent=styles["Normal"], fontSize=8,   textColor=LOW_FG)),
        Paragraph("Low Stock (&lt;500 kg)", ParagraphStyle("ollt", parent=styles["Normal"], fontSize=7.5, textColor=LOW_FG)),
    ]]
    legend = Table(legend_data, colWidths=[5 * mm, 30 * mm, 5 * mm, 40 * mm])
    legend.setStyle(TableStyle([
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    # ── Assemble story ────────────────────────────────────────────────────────
    story = [
        summary_tbl,
        Spacer(1, 6 * mm),
    ]
 
    for sec in normalised_sections:
        story.append(Paragraph(sec["title"], section_title_style))
        story.append(HRFlowable(width="100%", thickness=0.6, color=RULE, spaceAfter=4))
        story.append(build_section_table(sec["headers"], sec["rows"]))
        story.append(Spacer(1, 6 * mm))
 
    # Replace the last spacer with a tighter gap before the legend
    if story and isinstance(story[-1], Spacer):
        story[-1] = Spacer(1, 3 * mm)
 
    story.append(legend)
 
    # ── Build PDF ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN_H, rightMargin=MARGIN_H,
        topMargin=MARGIN_TOP, bottomMargin=MARGIN_BOT,
    )
    doc.build(story, onFirstPage=decorator, onLaterPages=decorator)
    buffer.seek(0)
    return buffer.read()

# Handle export dispatch report excel.

def export_dispatch_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatch Report"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Dispatch Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = _excel_fill("0F2A44")

    ws.append([f"Generated: {_generated_at_text()} | Records: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="334155", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E8F0FA")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("1F4E79"),
        alt_row_fill=_excel_fill("F4F8FC"),
        border_color="9CB3C7",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export dispatch entry report excel.

def export_dispatch_entry_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatch Entry"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Dispatch Entry Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = _excel_fill("115E59")

    ws.append([f"Generated: {_generated_at_text()} | Rows: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="134E4A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E6F6F4")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("0F766E"),
        alt_row_fill=_excel_fill("F1FBFA"),
        border_color="8AB8B3",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export raw material report excel.

def export_raw_material_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Material"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Raw Material Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = _excel_fill("5A3A12")

    ws.append([f"Generated: {_generated_at_text()} | Entries: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="6B4F2A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("F7EEDF")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("7C4A03"),
        alt_row_fill=_excel_fill("FCF7EF"),
        border_color="C6A57A",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export raw material entry report excel.

def export_raw_material_entry_report_excel(sections: list[dict]) -> bytes:
    wb = Workbook()
    section_list = sections or [{"title": "Data", "headers": ["Data"], "rows": []}]

    for idx, section in enumerate(section_list):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = str(section.get("sheet_name") or section.get("title") or f"Section{idx + 1}")[:31]

        title = str(section.get("title") or "Section")
        section_headers = section.get("headers") or ["Data"]
        section_rows = _normalize_rows(section.get("rows") or [])
        if not section_rows:
            section_rows = [["No data available"]]

        col_count = max(len(section_headers), 1)

        ws.append(["Raw Material Entry Report"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = _excel_fill("7C2D12")

        ws.append([f"{title} | Generated: {_generated_at_text()}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws["A2"].font = Font(italic=True, color="7C2D12", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].fill = _excel_fill("FFEDD5")

        ws.append([])
        ws.append(section_headers)
        for row in section_rows:
            ws.append([row[col] if col < col_count else "" for col in range(col_count)])

        header_row = 4
        first_data_row = 5
        last_data_row = max(first_data_row, ws.max_row)
        _excel_apply_table_style(
            ws,
            header_row=header_row,
            first_data_row=first_data_row,
            last_data_row=last_data_row,
            col_count=col_count,
            header_fill=_excel_fill("C2410C"),
            alt_row_fill=_excel_fill("FFF7ED"),
            border_color="E3B28A",
        )
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
        _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export production report excel.

def export_production_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Production"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Production Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("14532D")

    ws.append([f"Generated: {_generated_at_text()} | Batches: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="1F5135", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E7F5EA")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("166534"),
        alt_row_fill=_excel_fill("F2FBF4"),
        border_color="97B99E",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export batch report excel.

def export_batch_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Report"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Batch Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("1E3A8A")

    ws.append([f"Generated: {_generated_at_text()} | Records: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="1E3A8A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E8EEFF")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("1D4ED8"),
        alt_row_fill=_excel_fill("F3F7FF"),
        border_color="9FB6E7",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export batch consumption report excel.

def export_batch_consumption_report_excel(sections: list[dict]) -> bytes:
    wb = Workbook()
    section_list = sections or [{"title": "Data", "headers": ["Data"], "rows": []}]

    for idx, section in enumerate(section_list):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = str(section.get("sheet_name") or section.get("title") or f"Section{idx + 1}")[:31]

        title = str(section.get("title") or "Section")
        section_headers = section.get("headers") or ["Data"]
        section_rows = _normalize_rows(section.get("rows") or [])
        if not section_rows:
            section_rows = [["No data available"]]

        col_count = max(len(section_headers), 1)

        ws.append(["Batch Consumption Report"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = _excel_fill("1E3A8A")

        ws.append([f"{title} | Generated: {_generated_at_text()}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws["A2"].font = Font(italic=True, color="1E3A8A", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].fill = _excel_fill("E8EEFF")

        ws.append([])
        ws.append(section_headers)
        for row in section_rows:
            ws.append([row[col] if col < col_count else "" for col in range(col_count)])

        header_row = 4
        first_data_row = 5
        last_data_row = max(first_data_row, ws.max_row)
        _excel_apply_table_style(
            ws,
            header_row=header_row,
            first_data_row=first_data_row,
            last_data_row=last_data_row,
            col_count=col_count,
            header_fill=_excel_fill("1D4ED8"),
            alt_row_fill=_excel_fill("F3F7FF"),
            border_color="9FB6E7",
        )
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
        _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export rm stock report excel.

def export_rm_stock_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RM Stock"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["RM Stock Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("115E59")

    ws.append([f"Generated: {_generated_at_text()} | Rows: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="134E4A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E6F8F6")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("0F766E"),
        alt_row_fill=_excel_fill("F2FCFA"),
        border_color="8CBDB7",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export rm individual stock report excel.

def export_rm_individual_stock_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "RM Individual"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Individual Raw Material Stock"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("1F2937")

    ws.append([f"Generated: {_generated_at_text()} | Items: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="334155", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("EEF2F7")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("334155"),
        alt_row_fill=_excel_fill("F8FAFC"),
        border_color="B6BDC8",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export feed individual stock report excel.

def export_feed_individual_stock_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feed Individual"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or ["Feed Type", "Bag Size Mix", "Available Stock (kg)"]

    ws.append(["Individual Feed Available Stock"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("1E3A8A")

    ws.append([f"Generated: {_generated_at_text()} | Items: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="1E3A8A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("EAF0FF")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("1E3A8A"),
        alt_row_fill=_excel_fill("F5F8FF"),
        border_color="A6B6DF",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export feed stock report excel.

def export_feed_stock_report_excel(headers: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feed Stock"

    normalized_rows = _normalize_rows(rows)
    col_count = max(len(headers), 1)
    table_headers = headers or [f"Column {i + 1}" for i in range(col_count)]

    ws.append(["Feed Stock Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = _excel_fill("1E3A8A")

    ws.append([f"Generated: {_generated_at_text()} | Rows: {len(normalized_rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"].font = Font(italic=True, color="1E3A8A", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = _excel_fill("E9EFFF")

    ws.append([])
    ws.append(table_headers)
    for row in normalized_rows:
        ws.append([row[idx] if idx < col_count else "" for idx in range(col_count)])

    header_row = 4
    first_data_row = 5
    last_data_row = max(first_data_row, ws.max_row)
    _excel_apply_table_style(
        ws,
        header_row=header_row,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        col_count=col_count,
        header_fill=_excel_fill("1D4ED8"),
        alt_row_fill=_excel_fill("F2F7FF"),
        border_color="9DB4E6",
    )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
    _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export overall stock report excel.

def export_overall_stock_report_excel(sections: list[dict]) -> bytes:
    wb = Workbook()
    section_list = sections or [{"title": "Data", "headers": ["Data"], "rows": []}]

    for idx, section in enumerate(section_list):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = str(section.get("sheet_name") or section.get("title") or f"Section{idx + 1}")[:31]

        title = str(section.get("title") or "Section")
        section_headers = section.get("headers") or ["Data"]
        section_rows = _normalize_rows(section.get("rows") or [])
        if not section_rows:
            section_rows = [["No data available"]]

        col_count = max(len(section_headers), 1)

        ws.append(["Overall Stock Report"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = _excel_fill("111827")

        ws.append([f"{title} | Generated: {_generated_at_text()}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws["A2"].font = Font(italic=True, color="0F172A", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].fill = _excel_fill("EAEFF6")

        ws.append([])
        ws.append(section_headers)
        for row in section_rows:
            ws.append([row[col] if col < col_count else "" for col in range(col_count)])

        header_row = 4
        first_data_row = 5
        last_data_row = max(first_data_row, ws.max_row)
        _excel_apply_table_style(
            ws,
            header_row=header_row,
            first_data_row=first_data_row,
            last_data_row=last_data_row,
            col_count=col_count,
            header_fill=_excel_fill("0F172A"),
            alt_row_fill=_excel_fill("F7F9FC"),
            border_color="A7B1C2",
        )
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{ws.cell(row=header_row, column=col_count).coordinate[:-len(str(header_row))]}{last_data_row}"
        _excel_autofit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

# Handle export table to csv.

def export_table_to_csv(headers: list, rows: list) -> str:
    import csv

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue()
 
# Handle export batch report pdf.

def export_batch_report_pdf(
    batch,
    report,
    materials,
    plc_rows=None,
    plc_start=None,
    plc_end=None,
):
    # ── Palette ────────────────────────────────────────────────────────────────
    NAVY       = HexColor("#0D2545")
    NAVY_LIGHT = HexColor("#1A3A6B")
    GOLD       = HexColor("#C8922A")
    GOLD_LIGHT = HexColor("#F5E6C8")
    GREY_DARK  = HexColor("#3D3D3D")
    GREY_MID   = HexColor("#6B6B6B")
    GREY_LIGHT = HexColor("#F2F4F7")
    WHITE      = HexColor("#FFFFFF")
    RULE       = HexColor("#D9DDE6")
 
    GC = {
        "temp":        "#E05555",
        "humidity":    "#22BCD4",
        "conditioner": "#1DA462",
        "bagging":     "#F5964A",
        "speed":       "#9B59B6",
        "load":        "#1DA462",
        "p_before":    "#1DA462",
        "p_after":     "#D4A017",
    }
 
    IST_TZ      = timezone(timedelta(hours=5, minutes=30))
    generated_at = datetime.now(IST_TZ).strftime("%d %B %Y, %H:%M")
    company_name = getattr(batch, "company_name", "POULTRY NET")
    batch_start  = plc_start or getattr(batch, "hmi_started_at",   None)
    batch_end    = plc_end   or getattr(batch, "hmi_completed_at", None)
 
    # ── Inner helpers ──────────────────────────────────────────────────────────
 
    def _safe(v):
        if v in (None, ""):
            return "-"
        if isinstance(v, float):
            t = f"{v:.3f}".rstrip("0").rstrip(".")
            return t if t else "0"
        return str(v)
 
    # Handle fmt date.

    def _fmt_date(value):
        if value is None:
            return "-"
        if isinstance(value, datetime):
            utc = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
            return utc.astimezone(IST_TZ).strftime("%d %b %Y")
        return str(value)
 
    # Handle fmt ts.

    def _fmt_ts(value):
        if value is None:
            return "-"
        if isinstance(value, datetime):
            utc = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
            return utc.astimezone(IST_TZ).strftime("%d %b %Y %I:%M:%S %p IST")
        return str(value)
 
    # Handle panel.

    def _panel(title, content, width):
        ts = ParagraphStyle("PT", fontSize=9, fontName="Helvetica-Bold",
                             textColor=WHITE, alignment=TA_CENTER)
        box = Table([[Paragraph(title, ts)], [content]], colWidths=[width])
        box.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, 0), NAVY),
            ("LINEBELOW",     (0, 0), (0, 0), 1.5, GOLD),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
            ("TOPPADDING",    (0, 0), (0, 0), 7),
            ("BOTTOMPADDING", (0, 0), (0, 0), 7),
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ]))
        return box
 
    # Handle data table.

    def _data_table(headers, rows, width, col_ratios=None):
        n  = len(headers)
        cw = [width * r for r in (col_ratios or [1 / n] * n)]
        hs = ParagraphStyle("TH", fontSize=8, fontName="Helvetica-Bold",
                             textColor=WHITE, alignment=TA_CENTER)
        cs = ParagraphStyle("TD", fontSize=8, leading=10,
                             textColor=GREY_DARK, alignment=TA_LEFT)
        vs = ParagraphStyle("TV", fontSize=8, leading=10,
                             textColor=NAVY_LIGHT, alignment=TA_LEFT)
        data = [[Paragraph(h, hs) for h in headers]]
        for row in rows:
            data.append([Paragraph(_safe(c), vs if j == len(row) - 1 else cs)
                         for j, c in enumerate(row)])
        tbl = Table(data, colWidths=cw)
        cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.5, GOLD),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
            ("INNERGRID",     (0, 0), (-1, -1), 0.4, RULE),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWHEIGHT",     (0, 1), (-1, -1), 18),
        ]
        for i in range(1, len(data)):
            cmds.append(("BACKGROUND", (0, i), (-1, i),
                         GOLD_LIGHT if i % 2 == 0 else WHITE))
        tbl.setStyle(TableStyle(cmds))
        return tbl
 
    # Handle page decorator.

    def _page_decorator(canvas, doc):
        W, H = A4
        canvas.saveState()
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 28 * mm, W, 28 * mm, fill=True, stroke=False)
        canvas.setFillColor(GOLD)
        canvas.rect(0, H - 29.5 * mm, W, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(18 * mm, H - 12 * mm, company_name.upper())
        canvas.setFillColor(GOLD_LIGHT)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(18 * mm, H - 20 * mm, "PRODUCTION BATCH REPORT")
        meta_x = W - 18 * mm
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(meta_x, H - 11 * mm, f"Generated: {generated_at}")
        canvas.setFillColor(NAVY_LIGHT)
        canvas.roundRect(W - 36 * mm, H - 27 * mm, 16 * mm, 6 * mm, 1.5 * mm, fill=True, stroke=False)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawCentredString(W - 28 * mm, H - 24.2 * mm, f"Page {doc.page}")
        canvas.setFillColor(GREY_LIGHT)
        canvas.rect(0, 0, W, 14 * mm, fill=True, stroke=False)
        canvas.setFillColor(RULE)
        canvas.rect(0, 14 * mm, W, 0.4 * mm, fill=True, stroke=False)
        canvas.setFillColor(GREY_MID)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(18 * mm, 5.5 * mm,
            f"© {datetime.now().year} {company_name}  |  Confidential – Internal Use Only")
        canvas.drawRightString(W - 18 * mm, 5.5 * mm, "Production Management System")
        canvas.restoreState()
 
    # ── Graph helpers (nested) ─────────────────────────────────────────────────
 
    def _apply_rc():
        plt.rcParams.update({
            "figure.facecolor":   "#FFFFFF",
            "axes.facecolor":     "#FAFCFF",
            "axes.edgecolor":     "#D9DDE6",
            "axes.linewidth":     0.8,
            "axes.grid":          True,
            "grid.color":         "#E4EAF2",
            "grid.linewidth":     0.5,
            "grid.linestyle":     "-",
            "xtick.labelsize":    6,
            "ytick.labelsize":    6,
            "xtick.color":        "#6B6B6B",
            "ytick.color":        "#6B6B6B",
            "axes.labelsize":     6.5,
            "axes.labelcolor":    "#3D3D3D",
            "axes.titlesize":     8.5,
            "axes.titleweight":   "bold",
            "axes.titlecolor":    "#0D2545",
            "axes.titlelocation": "left",
            "legend.fontsize":    6,
            "legend.frameon":     True,
            "legend.framealpha":  0.9,
            "legend.edgecolor":   "#D9DDE6",
            "lines.linewidth":    1.2,
            "lines.markersize":   2,
            "lines.marker":       "o",
        })
 
    # Handle time labels.

    def _time_labels(rows, n):
        if rows:
            labels = []
            for i, row in enumerate(rows):
                ts = getattr(row, "recorded_at", None)
                if isinstance(ts, datetime):
                    utc = ts.replace(tzinfo=timezone.utc) if ts.tzinfo is None else ts.astimezone(timezone.utc)
                    labels.append(utc.astimezone(IST_TZ).strftime("%H:%M:%S"))
                elif ts not in (None, ""):
                    labels.append(str(ts))
                else:
                    labels.append(str(i + 1))
            return labels

        base = datetime(2024, 1, 1, 11, 2, 35)
        return [(base + timedelta(seconds=i * 15)).strftime("%H:%M:%S") for i in range(n)]
 
    # Set xticks.

    def _set_xticks(ax, xs, labels):
        step = max(1, len(xs) // 8)
        idx  = list(range(0, len(xs), step))
        ax.set_xticks([xs[i] for i in idx])
        ax.set_xticklabels([labels[i] for i in idx], rotation=0, ha="center")
        ax.set_xlabel("Time", labelpad=3)
 
    # Handle avg line.

    def _avg_line(ax, data, color):
        avg = float(np.mean(data))
        ax.axhline(avg, color=color, linestyle="--", linewidth=1.0, alpha=0.8, zorder=4)
        return avg
 
    # Handle fill.

    def _fill(ax, xs, data, color):
        ax.fill_between(xs, data, alpha=0.08, color=color, zorder=1)
 
    # Save value.

    def _save(fig):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        return buf
 
    # Handle img.

    def _img(buf, w, h):
        buf.seek(0)
        return Image(buf, width=w, height=h)
 
    # Graph 1 – Temperature Trends  (Y: 0–100 step 10)

    def _graph_temperature(rows, w_pt, h_pt):
        n  = len(rows) if rows else 100
        xs = np.arange(n)
        tl = _time_labels(rows, n)
        if rows:
            amb  = np.array([getattr(r, "ambient_temp", 30) for r in rows], dtype=float)
            hum  = np.array([getattr(r, "humidity", 59) for r in rows], dtype=float)
            cond = np.array([getattr(r, "conditioner_temp", 84) for r in rows], dtype=float)
            bag  = np.array([getattr(r, "bagging_temp", 33) for r in rows], dtype=float)
        else:
            amb  = np.clip(30 + np.cumsum(np.random.randn(n) * 0.3) * 0.15, 22, 40)
            hum  = np.clip(59 + np.cumsum(np.random.randn(n) * 0.4) * 0.15, 40, 80)
            cond = np.clip(84 + np.cumsum(np.random.randn(n) * 0.2) * 0.10, 70, 95)
            bag  = np.clip(33 + np.cumsum(np.random.randn(n) * 0.3) * 0.10, 25, 45)
 
        _apply_rc()
        fig, ax = plt.subplots(figsize=(w_pt / 72, h_pt / 72))
        ax.plot(xs, amb,  color=GC["temp"],        label="Ambient Temp",     zorder=3, markevery=0.07)
        ax.plot(xs, hum,  color=GC["humidity"],    label="Humidity",          zorder=3, markevery=0.07)
        ax.plot(xs, cond, color=GC["conditioner"], label="Conditioner Temp",  zorder=3, markevery=0.07)
        ax.plot(xs, bag,  color=GC["bagging"],     label="Bagging Temp",      zorder=3, markevery=0.07)
        for data, c in [(amb, GC["temp"]), (hum, GC["humidity"]),
                        (cond, GC["conditioner"]), (bag, GC["bagging"])]:
            _fill(ax, xs, data, c)
        avg_amb  = _avg_line(ax, amb,  GC["temp"])
        avg_hum  = _avg_line(ax, hum,  GC["humidity"])
        avg_cond = _avg_line(ax, cond, GC["conditioner"])
        avg_bag  = _avg_line(ax, bag,  GC["bagging"])
        ax.set_ylim(0, 100)
        ax.yaxis.set_major_locator(mticker.MultipleLocator(10))
        ax.set_ylabel("Temperature (°C) / Humidity (%)")
        ax.set_title("Temperature Trends")
        _set_xticks(ax, xs, tl)
        ax.text(0.99, 1.025,
                f"Avg A: {avg_amb:.2f}  |  Avg H: {avg_hum:.2f}  |  Avg C: {avg_cond:.2f}  |  Avg B: {avg_bag:.2f}",
                transform=ax.transAxes, fontsize=6, ha="right", va="bottom",
                color=GC["temp"], fontweight="bold", clip_on=False)
        ax.legend(loc="lower center", ncol=4, bbox_to_anchor=(0.5, -0.22), markerscale=0.8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout(pad=0.6)
        return _save(fig)
 
    # Graph 2 – Pressure Trends  (Y: 0–5 step 1)

    def _graph_pressure(rows, w_pt, h_pt):
        n  = len(rows) if rows else 100
        xs = np.arange(n)
        tl = _time_labels(rows, n)
        if rows:
            pb = np.array([getattr(r, "pressure_before", 3.90) for r in rows], dtype=float)
            pa = np.array([getattr(r, "pressure_after", 2.95) for r in rows], dtype=float)
        else:
            pb = np.clip(3.90 + np.cumsum(np.random.randn(n) * 0.03) * 0.06, 0, 5)
            pa = np.clip(2.95 + np.cumsum(np.random.randn(n) * 0.02) * 0.04, 0, 5)
 
        _apply_rc()
        fig, ax = plt.subplots(figsize=(w_pt / 72, h_pt / 72))
        ax.plot(xs, pb, color=GC["p_before"], label="Pressure Before (bar)", zorder=3, markevery=0.07)
        ax.plot(xs, pa, color=GC["p_after"],  label="Pressure After (bar)",  zorder=3, markevery=0.07)
        _fill(ax, xs, pb, GC["p_before"])
        _fill(ax, xs, pa, GC["p_after"])
        avg_pb = _avg_line(ax, pb, GC["p_before"])
        avg_pa = _avg_line(ax, pa, GC["p_after"])
        ax.set_ylim(0, 20)
        ax.yaxis.set_major_locator(mticker.MultipleLocator(2))
        ax.set_ylabel("Pressure (bar)")
        ax.set_title("Pressure Trends")
        _set_xticks(ax, xs, tl)
        ax.text(0.99, 1.025, f"Avg P1: {avg_pb:.2f}  |  Avg P2: {avg_pa:.2f}",
                transform=ax.transAxes, fontsize=6, ha="right", va="bottom",
                color=GC["p_before"], fontweight="bold", clip_on=False)
        ax.legend(loc="lower center", ncol=2, bbox_to_anchor=(0.5, -0.22), markerscale=0.8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout(pad=0.6)
        return _save(fig)
 
    # Graph 3 – Feeder Speed vs Load  (Left Y: 0–1500 step 150 | Right Y: 0–300 step 30)

    def _graph_feeder(rows, w_pt, h_pt):
        n  = len(rows) if rows else 100
        xs = np.arange(n)
        tl = _time_labels(rows, n)
        if rows:
            spd = np.array([getattr(r, "pellet_feeder_speed", 120) for r in rows], dtype=float)
            ld  = np.array([getattr(r, "pellet_motor_load", 73) for r in rows], dtype=float)
        else:
            spd = np.clip(120 + np.cumsum(np.random.randn(n) * 0.5) * 0.3,  80, 200)
            ld  = np.clip( 73 + np.cumsum(np.random.randn(n) * 0.8) * 0.2,  30, 120)
 
        c_spd = GC["speed"]
        c_ld  = GC["load"]
 
        _apply_rc()
        fig, ax1 = plt.subplots(figsize=(w_pt / 72, h_pt / 72))
 
        ax1.plot(xs, spd, color=c_spd, label="Pellet Feeder Speed (rpm)", zorder=3, markevery=0.07)
        _fill(ax1, xs, spd, c_spd)
        avg_spd = _avg_line(ax1, spd, c_spd)
        ax1.set_ylim(0, 1500)
        ax1.yaxis.set_major_locator(mticker.MultipleLocator(150))
        ax1.set_ylabel("Pellet Feeder Speed (rpm)", color=c_spd, fontsize=6.5)
        ax1.tick_params(axis="y", labelcolor=c_spd, labelsize=6)
        ax1.spines["left"].set_color(c_spd)
 
        ax2 = ax1.twinx()
        ax2.plot(xs, ld, color=c_ld, label="Pellet Feeder Load (amp)", zorder=2, markevery=0.07)
        _fill(ax2, xs, ld, c_ld)
        avg_ld = _avg_line(ax2, ld, c_ld)
        ax2.set_ylim(0, 300)
        ax2.yaxis.set_major_locator(mticker.MultipleLocator(30))
        ax2.set_ylabel("Pellet Feeder Load (amp)", color=c_ld, fontsize=6.5)
        ax2.tick_params(axis="y", labelcolor=c_ld, labelsize=6)
        ax2.spines["right"].set_color(c_ld)
 
        ax1.set_title("Pellet Feeder Speed vs Load")
        _set_xticks(ax1, xs, tl)
        ax1.text(0.99, 1.025, f"Avg RPM: {avg_spd:.2f}  |  Avg Amp: {avg_ld:.2f}",
                 transform=ax1.transAxes, fontsize=6, ha="right", va="bottom",
                 color=c_spd, fontweight="bold", clip_on=False)
        h1, l1 = ax1.get_legend_handles_labels()
        h2, l2 = ax2.get_legend_handles_labels()
        ax1.legend(h1 + h2, l1 + l2, loc="lower center", ncol=2,
                   bbox_to_anchor=(0.5, -0.22), markerscale=0.8)
        ax1.spines["top"].set_visible(False)
        ax2.spines["top"].set_visible(False)
        fig.tight_layout(pad=0.6)
        return _save(fig)
 
    # ── Build document ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=36 * mm,  bottomMargin=20 * mm,
    )
    W      = doc.width
    story  = []
    styles = getSampleStyleSheet()
 
    # Summary strip
    slbl = ParagraphStyle("SL", fontSize=8, fontName="Helvetica-Bold", textColor=NAVY)
    sval = ParagraphStyle("SV", fontSize=8, textColor=GREY_MID)
    summ = Table(
        [[Paragraph("BATCH SUMMARY", slbl),
          Paragraph(f"Batch ID: <b>{_safe(batch.id)}</b>", sval),
          Paragraph(f"Product: <b>{_safe(batch.product_name)}</b>", sval),
          Paragraph(f"Generated: <b>{generated_at}</b>", sval)]],
        colWidths=[W * w for w in (0.22, 0.20, 0.28, 0.30)],
    )
    summ.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GOLD_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, GOLD),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(summ)
    story.append(Spacer(1, 7 * mm))
 
    # Batch detail cells
    lbl_s = ParagraphStyle("BL", fontSize=7.5, fontName="Helvetica-Bold",
                            textColor=GREY_MID, leading=10)
    val_s = ParagraphStyle("BV", fontSize=9, fontName="Helvetica-Bold",
                            textColor=NAVY, leading=12)
 
    # Handle bcell.

    def _bcell(label, value):
        return Table(
            [[Paragraph(label, lbl_s)], [Paragraph(_safe(value), val_s)]],
            colWidths=[W / 5 - 4],
        )
 
    items  = [
        ("Batch ID",    batch.id),
        ("Date",        _fmt_date(batch.date)),
        ("Product",     batch.product_name),
        ("Batch Size",  batch.batch_size),
        ("Output",      batch.output),
        ("Start Time",  _fmt_ts(batch_start)),
        ("End Time",    _fmt_ts(batch_end)),
        ("No. of Bags", getattr(batch, "num_bags",       "") or ""),
        ("Weight/Bag",  getattr(batch, "weight_per_bag", "") or ""),
        ("Water",       getattr(batch, "water",          "") or ""),
    ]
    bcells = [_bcell(l, v) for l, v in items]
    batch_inner = Table(
        [bcells[:5], bcells[5:]],
        colWidths=[W / 5] * 5,
        rowHeights=[50, 50],
    )
    batch_inner.setStyle(TableStyle([
        ("VALIGN",         (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [GOLD_LIGHT, WHITE]),
        ("BOX",            (0, 0), (-1, -1), 0.6, RULE),
        ("INNERGRID",      (0, 0), (-1, -1), 0.4, RULE),
        ("LINEBELOW",      (0, 0), (-1, 0),  1.5, GOLD),
        ("LEFTPADDING",    (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
        ("TOPPADDING",     (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 8),
    ]))
    story.append(_panel("Batch Summary", batch_inner, W))
    story.append(Spacer(1, 7 * mm))
 
    # Raw materials (show actual consumed quantity where available).

    def _material_name(row):
        if isinstance(row, dict):
            return row.get("rm_name", "")
        return getattr(row, "rm_name", "")

    # Handle material quantity.

    def _material_quantity(row):
        if isinstance(row, dict):
            if row.get("_report_quantity") is not None:
                return row.get("_report_quantity")
            if row.get("total_quantity") is not None:
                return row.get("total_quantity")
            return row.get("quantity")
        report_quantity = getattr(row, "_report_quantity", None)
        if report_quantity is not None:
            return report_quantity
        total_quantity = getattr(row, "total_quantity", None)
        if total_quantity is not None:
            return total_quantity
        return getattr(row, "quantity", None)

    mat_rows = [[_material_name(m), _material_quantity(m)] for m in materials] or [["No materials recorded", "-"]]
    story.append(_panel("Raw Material Consumption",
                         _data_table(["Material", "Consumed Quantity (kg)"], mat_rows, W,
                                     col_ratios=[0.65, 0.35]), W))
    story.append(Spacer(1, 7 * mm))
 
    # Chemical & Physical side-by-side
    half      = (W - 8) / 2
    chem_rows = [
        ["Protein",    report.protein],
        ["Fat",        report.fat],
        ["Fiber",      report.fiber],
        ["Ash",        report.ash],
        ["Calcium",    report.calcium],
        ["Phosphorus", report.phosphorus],
        ["Salt",       report.salt],
    ]
    phys_rows = [
        ["HM Retention",         report.hm_retention],
        ["Mixer Moisture",       report.mixer_moisture],
        ["Conditioner Moisture", report.conditioner_moisture],
        ["Moisture Addition",    report.moisture_addition],
        ["Final Feed Moisture",  report.final_feed_moisture],
        ["Water Activity",       report.water_activity],
        ["Hardness",             report.hardness],
        ["Pellet Diameter",      report.pellet_diameter],
        ["Fines",                report.fines],
    ]
    side = Table(
        [[_panel("Chemical Analysis",
                  _data_table(["Chemical Parameter", "Value"], chem_rows, half,
                               col_ratios=[0.62, 0.38]), half),
          _panel("Physical Parameters",
                  _data_table(["Physical Parameter", "Value"], phys_rows, half,
                               col_ratios=[0.62, 0.38]), half)]],
        colWidths=[half + 4, half + 4],
    )
    side.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(side)
    story.append(PageBreak())
 
    # Sensor graphs — 3 stacked full-width
    rows  = plc_rows or []
    GH    = 185
    GH_FD = 195
 
    graph_tbl = Table(
        [[_img(_graph_temperature(rows, W, GH),    W, GH)],
         [_img(_graph_pressure(rows,    W, GH),    W, GH)],
         [_img(_graph_feeder(rows,      W, GH_FD), W, GH_FD)]],
        colWidths=[W],
    )
    graph_tbl.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 2),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BOX",           (0, 0), (-1, -1), 0.6, RULE),
        ("INNERGRID",     (0, 0), (-1, -1), 0.4, RULE),
    ]))
    story.append(_panel("Sensor Trends", graph_tbl, W))
 
    doc.build(story, onFirstPage=_page_decorator, onLaterPages=_page_decorator)
    buffer.seek(0)
    return buffer.read()
 
